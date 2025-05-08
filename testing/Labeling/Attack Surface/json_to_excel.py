"""
This script is used to convert the parsed results and BERT classification data into an 
Excel file for manual labeling. Use this script if you want to create an excel file for labeling 
a new project. The BERT classifications are included as a guide to help the user label the data.

NOTE: You must first run the project you want with the tool to generate the parsedResults.json and Data.json.
"""

import json
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill
import math

# Predefined Sink Categories
SINK_CATEGORIES = [
    "N/A", "I/O Sink", "Print Sink", "Network Sink", "Log Sink", 
    "Database Sink", "Email Sink", "IPC Sink", "Clipboard Sink",
    "GUI Display Sink", "RPC Sink", "Environment Variable Sink",
    "Command Execution Sink", "Configuration File Sink"
]

# Predefined Yes/No options for Kyler and Sara
YES_NO_OPTIONS = ["Yes", "No"]

# Function to read JSON files
def read_json(file_path):
    with open(file_path, 'r', encoding='UTF-8') as file:
        return json.load(file)

# Function to restructure BERT data with 'fileName' as the key
def restructure_bert_data(bert_data):
    bert_data_dict = {}
    for item in bert_data:
        file_name = item['fileName']
        bert_data_dict[file_name] = item
    return bert_data_dict

# Function to classify variables, strings, comments, and sinks as "Yes" or "No"
def classify_bert_vs_parsed(parsed_item, bert_item):
    output = {}
    for key in ['variables', 'strings', 'comments', 'sinks']:
        parsed_raw = parsed_item.get(key, [])
        # If items are dictionaries, extract the 'name' value
        if parsed_raw and isinstance(parsed_raw[0], dict):
            parsed_values = {item.get('name') for item in parsed_raw}
        else:
            parsed_values = set(parsed_raw)
        bert_values = {item['name'] for item in bert_item.get(key, [])}
        classified_values = []
        for value in parsed_values:
            classified_values.append((value, "Yes" if value in bert_values else "No"))
        output[key] = classified_values
    return output

# Function to check if a file has any relevant data in all types
def has_relevant_data(output):
    for values in output.values():
        if values:  # If there are any classified items, the file has relevant data
            return True
    return False

def load_file_paths(src_dir, project_dir):
    file_paths = {}
    for root, dirs, files in os.walk(src_dir):
        for file in files:
            file_paths[file] = os.path.relpath(os.path.join(root, file), project_dir)
    return file_paths

# Replace find_file_in_src function with a lookup
def find_file_in_lookup(file_paths, file_name):
    return file_paths.get(file_name, None)

# Function to format the output for Excel
def format_for_excel(output_data, src_dir, file_paths):
    formatted_rows = []

    for file_name, data in output_data.items():
        # Find the file path recursively in the src directory
        result = find_file_in_lookup(file_paths, file_name)
        if result:
            file_path = result
        else:
            print(f"Warning: File '{file_name}' not found in {src_dir}")
            continue  # Skip if file not found


        # Add the file name and file path as a header row
        formatted_rows.append([f"{file_name}", f"{file_path}", "", "", ""])

        # Add category headers
        formatted_rows.append(["Category", "Item", "BERT Classification", "Kyler", "Sara"])

        # Ensure that all categories are written, even if they are empty
        for category in ['variables', 'strings', 'comments', 'sinks']:
            classified_items = data.get(category, [])

            # Add category name as a row
            formatted_rows.append([category.capitalize(), "", "", ""])


            # If there are no items for this category, add a blank row for spacing
            if not classified_items:
                formatted_rows.append(["", "", "", "", ""])

            # If no classified items, leave the section blank but keep the header
            else:
                # Write each classified item with its classification (Yes/No)
                if category == 'sinks':  # For sinks, add 'N/A' for the category column
                    for item, classification in classified_items:
                        formatted_rows.append(["N/A", item, classification, "", ""])
                else:
                    for item, classification in classified_items:
                        formatted_rows.append(["", item, classification, "", ""])

    return formatted_rows

from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils.cell import coordinate_from_string
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.cell_range import MultiCellRange
from openpyxl.styles import PatternFill
import math

def group_cells(cell_list):
    """
    Given a list of cell coordinates (e.g., ["D2", "D3", "D5"]),
    group contiguous cells in the same column into a list of range strings.
    For example, if cell_list contains ["D2", "D3", "D4", "D6"],
    this returns ["D2:D4", "D6"].
    """
    grouped = {}
    for coord in cell_list:
        col, row = coordinate_from_string(coord)
        row = int(row)
        grouped.setdefault(col, []).append(row)
    ranges = []
    for col, rows in grouped.items():
        rows.sort()
        start = rows[0]
        end = rows[0]
        for r in rows[1:]:
            if r == end + 1:
                end = r
            else:
                if start == end:
                    ranges.append(f"{col}{start}")
                else:
                    ranges.append(f"{col}{start}:{col}{end}")
                start = r
                end = r
        if start == end:
            ranges.append(f"{col}{start}")
        else:
            ranges.append(f"{col}{start}:{col}{end}")
    return ranges

def finalize_workbook(output_file):
    wb = load_workbook(output_file)
    ws = wb.active

    # Create DataValidation objects with error messages and "stop" style.
    dv_sinks = DataValidation(
        type="list",
        formula1=f'"{",".join(SINK_CATEGORIES)}"',
        showDropDown=True,
        error="Please select one of the allowed sink categories.",
        errorTitle="Invalid Entry",
        errorStyle="stop"
    )
    dv_yes_no = DataValidation(
        type="list",
        formula1=f'"{",".join(YES_NO_OPTIONS)}"',
        showDropDown=True,
        allow_blank=True,
        error="Please select Yes or No.",
        errorTitle="Invalid Entry",
        errorStyle="stop"
    )

    total_rows = ws.max_row
    print(f"Total rows: {total_rows}")

    # Apply blue fill via conditional formatting for header rows.
    blue_fill = PatternFill(start_color="008cff", end_color="008cff", fill_type="solid")
    header_rule = FormulaRule(formula=['$A1="Category"'], fill=blue_fill)
    ws.conditional_formatting.add(f"A1:E{total_rows}", header_rule)

    # Collect cell addresses where we want the validations.
    sink_addresses = []
    yes_no_addresses = []

    for row in ws.iter_rows(min_row=1, max_row=total_rows, max_col=5):
        row_num = row[0].row
        first_val = row[0].value
        second_val = row[1].value

        # For sink validations, look in column A for "N/A"
        if first_val is not None and str(first_val).strip() == "N/A":
            sink_addresses.append(row[0].coordinate)

        # For Yes/No validations, target columns 4 and 5 when column B is non-empty
        # and the row is not a header (i.e. column A != "Category")
        if second_val and (first_val is None or str(first_val).strip() != "Category"):
            yes_no_addresses.append(ws.cell(row=row_num, column=4).coordinate)
            yes_no_addresses.append(ws.cell(row=row_num, column=5).coordinate)

        if row_num % 1000 == 0:
            current_progress = math.ceil(100 * row_num / total_rows)
            print(f"Progress: {current_progress}%")

    # Group contiguous addresses into ranges.
    sink_ranges = group_cells(sink_addresses)
    yes_no_ranges = group_cells(yes_no_addresses)
    print("Sink ranges:", sink_ranges)
    print("Yes/No ranges:", yes_no_ranges)

    # Build MultiCellRange objects from the list of range strings.
    if sink_ranges:
        dv_sinks.sqref = MultiCellRange(sink_ranges)
        ws.add_data_validation(dv_sinks)
    if yes_no_ranges:
        dv_yes_no.sqref = MultiCellRange(yes_no_ranges)
        ws.add_data_validation(dv_yes_no)

    # Protect the worksheet so that data validation is strictly enforced.
    # (Ensure that the cells with validations remain locked—by default, cells are locked.)
    ws.protection.sheet = True
    # Optionally set a password: ws.protection.password = "YourPassword"

    wb.save(output_file)


# Function to save the formatted output to an Excel file
def save_to_excel(output_data, output_file):
    # Convert to a DataFrame
    df = pd.DataFrame(output_data)

    # Write to an Excel file
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, header=False)



# Main logic
def main():
    # File paths for parsed and BERT classification files
    project_name = 'CWEToyDataset'
    project_dir = os.path.join('backend', 'Files', project_name)
    src_dir = os.path.join(project_dir, project_name)  # Assume the source files are inside 'src' directory
    
    parsed_file_path = os.path.join(project_dir, 'parsedResults.json')  
    bert_file_path = os.path.join(project_dir, 'data.json')  
    
    # Output Excel file
    output_file_path = os.path.join('testing', 'Labeling', 'Attack Surface', f'{project_name}.xlsx')
    
    # Read both JSON files
    parsed_data = read_json(parsed_file_path)
    bert_data = read_json(bert_file_path)

    # Restructure BERT data to use fileName as the key
    bert_data_dict = restructure_bert_data(bert_data)
    print("Data loaded successfully")
    
    # Dictionary to store the final output
    output_data = {}

    # Process each file in the parsed JSON
    for file_name, parsed_item in parsed_data.items():
        # Find corresponding BERT classification (if exists)
        bert_item = bert_data_dict.get(file_name, {})
        
        # Classify items
        classified_data = classify_bert_vs_parsed(parsed_item, bert_item)
        
        # Only include files that have at least one non-empty section
        if has_relevant_data(classified_data):
            output_data[file_name] = classified_data
    
    print("Data classified successfully")
    # Load file paths for the source directory
    file_paths = load_file_paths(src_dir, project_dir)
    print("File paths loaded successfully")


    # Format the output for Excel and include the file path
    excel_output = format_for_excel(output_data, src_dir, file_paths)
    print("Data formatted successfully")
    
    # Save the output to an Excel file
    save_to_excel(excel_output, output_file_path)
    print("Data saved to Excel successfully")
    
    # Finalize the workbook: apply validation, formatting, and highlighting in one pass
    finalize_workbook(output_file_path)
    print("Workbook finalized successfully")

    print(f"Results saved to {output_file_path}")

# Run the script
if __name__ == '__main__':
    main()
