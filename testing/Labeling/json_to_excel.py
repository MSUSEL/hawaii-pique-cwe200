import json
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill

# Predefined Sink Categories
SINK_CATEGORIES = [
    "N/A", "I/O Sink", "Print Sink", "Network Sink", "Log Sink", 
    "Database Sink", "Email Sink", "IPC Sink", "Clipboard Sink",
    "GUI Display Sink", "RPC Sink", "Environment Variable Sink",
    "Command Execution Sink", "Configuration File Sink"
]

# Predefined Yes/No options for Kyler and Sara
YES_NO_OPTIONS = ["Yes", "No"]

# Function to read JSON files
def read_json(file_path):
    with open(file_path, 'r', encoding='UTF-8') as file:
        return json.load(file)

# Function to restructure BERT data with 'fileName' as the key
def restructure_bert_data(bert_data):
    bert_data_dict = {}
    for item in bert_data:
        file_name = item['fileName']
        bert_data_dict[file_name] = item
    return bert_data_dict

# Function to classify variables, strings, comments, and sinks as "Yes" or "No"
def classify_bert_vs_parsed(parsed_item, bert_item):
    output = {}
    for key in ['variables', 'strings', 'comments', 'sinks']:
        parsed_values = set(parsed_item.get(key, []))
        bert_values = {item['name'] for item in bert_item.get(key, [])}
        classified_values = []
        for value in parsed_values:
            classified_values.append((value, "Yes" if value in bert_values else "No"))
        output[key] = classified_values
    return output

# Function to check if a file has any relevant data in all types
def has_relevant_data(output):
    for values in output.values():
        if values:  # If there are any classified items, the file has relevant data
            return True
    return False

def load_file_paths(src_dir, project_dir):
    file_paths = {}
    for root, dirs, files in os.walk(src_dir):
        for file in files:
            file_paths[file] = os.path.relpath(os.path.join(root, file), project_dir)
    return file_paths

# Replace find_file_in_src function with a lookup
def find_file_in_lookup(file_paths, file_name):
    return file_paths.get(file_name, None)

# Function to format the output for Excel
def format_for_excel(output_data, src_dir, file_paths):
    formatted_rows = []

    for file_name, data in output_data.items():
        # Find the file path recursively in the src directory
        if file_name not in file_paths:
            print(f"Warning: File '{file_name}' not found in {src_dir}")
            continue  # Skip if file not found
        else:
            file_path = file_paths[file_name]

        # Add the file name and file path as a header row
        formatted_rows.append([f"{file_name}", f"{file_path}", "", "", ""])

        # Add category headers
        formatted_rows.append(["Category", "Item", "BERT Classification", "Kyler", "Sara"])

        # Ensure that all categories are written, even if they are empty
        for category in ['variables', 'strings', 'comments', 'sinks']:
            classified_items = data.get(category, [])

            # Add category name as a row
            formatted_rows.append([category.capitalize(), "", "", ""])

            # If no classified items, leave the section blank but keep the header
            if classified_items:
                # Write each classified item with its classification (Yes/No)
                if category == 'sinks':  # For sinks, add 'N/A' for the category column
                    for item, classification in classified_items:
                        formatted_rows.append(["N/A", item, classification, "", ""])
                else:
                    for item, classification in classified_items:
                        formatted_rows.append(["", item, classification, "", ""])

    return formatted_rows

# Function to apply data validation, highlight headers, and save in one pass
def finalize_workbook(output_file):
    wb = load_workbook(output_file)
    ws = wb.active

    # Define drop-down lists
    dv_sinks = DataValidation(type="list", formula1=f'"{",".join(SINK_CATEGORIES)}"', showDropDown=True)
    dv_yes_no = DataValidation(type="list", formula1=f'"{",".join(YES_NO_OPTIONS)}"', showDropDown=True, allow_blank=True)

    # Define fill color for headers
    blue_fill = PatternFill(start_color="008cff", end_color="008cff", fill_type="solid")

    # Loop through all rows, applying formatting, validation, and headers in one pass
    for row in range(1, ws.max_row + 1):
        # Highlight headers
        if ws.cell(row=row, column=1).value == "Category":
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = blue_fill

        # Apply data validation to Sink Category
        if ws.cell(row=row, column=1).value == "N/A":
            dv_sinks.add(ws.cell(row, 1))

        # Apply Yes/No drop-down to Kyler and Sara columns, but skip the header rows
        if ws.cell(row=row, column=2).value and ws.cell(row=row, column=1).value != "Category":
            dv_yes_no.add(ws.cell(row, 4))  # Kyler column
            dv_yes_no.add(ws.cell(row, 5))  # Sara column

    # Add validations to the sheet and save in one pass
    ws.add_data_validation(dv_sinks)
    ws.add_data_validation(dv_yes_no)

    wb.save(output_file)

# Function to save the formatted output to an Excel file
def save_to_excel(output_data, output_file):
    # Convert to a DataFrame
    df = pd.DataFrame(output_data)

    # Write to an Excel file
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, header=False)



# Main logic
def main():
    # File paths for parsed and BERT classification files
    project_name = 'rocketmq-rocketmq-all-5.3.1'
    project_dir = os.path.join('backend', 'Files', project_name)
    src_dir = os.path.join(project_dir, project_name)  # Assume the source files are inside 'src' directory
    
    parsed_file_path = os.path.join(project_dir, 'parsedResults.json')  
    bert_file_path = os.path.join(project_dir, 'data.json')  
    
    # Output Excel file
    output_file_path = os.path.join('testing', 'Labeling', 'output_google_sheets.xlsx')
    
    # Read both JSON files
    parsed_data = read_json(parsed_file_path)
    bert_data = read_json(bert_file_path)

    # Restructure BERT data to use fileName as the key
    bert_data_dict = restructure_bert_data(bert_data)
    
    # Dictionary to store the final output
    output_data = {}

    # Process each file in the parsed JSON
    for file_name, parsed_item in parsed_data.items():
        # Find corresponding BERT classification (if exists)
        bert_item = bert_data_dict.get(file_name, {})
        
        # Classify items
        classified_data = classify_bert_vs_parsed(parsed_item, bert_item)
        
        # Only include files that have at least one non-empty section
        if has_relevant_data(classified_data):
            output_data[file_name] = classified_data
    
    # Load file paths for the source directory
    file_paths = load_file_paths(src_dir, project_dir)


    # Format the output for Excel and include the file path
    excel_output = format_for_excel(output_data, src_dir, file_paths)
    
    # Save the output to an Excel file
    save_to_excel(excel_output, output_file_path)
    
    # Finalize the workbook: apply validation, formatting, and highlighting in one pass
    finalize_workbook(output_file_path)

    print(f"Results saved to {output_file_path}")

# Run the script
if __name__ == '__main__':
    main()
