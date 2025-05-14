"""
This script takes in the projects from testing/Advisory/clean_advisories.txt
and attempts to create a CodeQL database for each project.
This is done by attempting to create a database for each project with 
Java 21, 17, 11, and 8.
If any of these attempts results in a status code of 0, the project is considered
to be buildable with that version of Java.

The metadata for each project is stored in a dictionary, containing 
repoName, projectVersion, javaVersion, projectName, and url.
The output is a JSON file containing the metadata for each project.
Along with a xlsx file.
"""

import os
import requests
import subprocess
import zipfile
import json
import time
import re
import platform
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import concurrent.futures
from datetime import datetime
from urllib.parse import unquote


input_projects = os.path.join("testing", "Advisory", "clean_advisories.txt")
java_versions = ["21", "8", "17", "11", ]
DOWNLOAD_DIR = os.path.join("testing", "PIQUE_Projects", "project_downloads")
DATABASE_DIR = os.path.join("testing", "PIQUE_Projects", "databases")
OUTPUT_DIR = os.path.join("testing", "PIQUE_Projects", "output")

GITHUB_TOKEN =  ''
GITHUB_HEADERS = {
    "Authorization": f"token {GITHUB_TOKEN}" if GITHUB_TOKEN else None,
    "Accept": "application/vnd.github+json"
}


def read_projects(file_path):
    """Reads the project list from a file."""
    with open(file_path, 'r') as f:
        projects = f.readlines()
    projects = [line.strip() for line in projects if line.strip()]
    return projects

def robust_github_get(url, headers=None, max_retries=5, timeout=10):
    for attempt in range(max_retries):
        try:
            response = requests.get(url, headers=headers, timeout=timeout)

            if response.status_code in [403, 429]:
                remaining = response.headers.get("X-RateLimit-Remaining")
                reset = response.headers.get("X-RateLimit-Reset")
                if remaining == "0" and reset:
                    sleep_time = int(reset) - int(time.time()) + 1
                    if sleep_time > 0:
                        print(f"Rate limit hit. Sleeping for {sleep_time} seconds...")
                        time.sleep(sleep_time)
                    continue

            if response.status_code == 404:
                return response  # Let caller handle

            if response.ok:
                return response

        except requests.exceptions.RequestException as e:
            print(f"[Attempt {attempt+1}] Request to {url} failed: {e}")
            time.sleep(2 ** attempt)

    print(f"[ERROR] Failed to fetch {url} after {max_retries} retries.")
    return None

def get_latest_tag_by_commit_date(owner, repo):
    tags_url = f"https://api.github.com/repos/{owner}/{repo}/tags"
    tags_response = robust_github_get(tags_url, headers=GITHUB_HEADERS)
    if tags_response.status_code != 200:
        return None

    tags = tags_response.json()
    latest_tag = None
    latest_date = None

    for tag in tags:
        commit_url = tag["commit"]["url"]
        commit_response = robust_github_get(commit_url, headers=GITHUB_HEADERS)
        if not commit_response or commit_response.status_code != 200:
            continue
        commit_data = commit_response.json()
        commit_date_str = commit_data.get("commit", {}).get("committer", {}).get("date")
        if not commit_date_str:
            continue
        commit_date = datetime.fromisoformat(commit_date_str.replace("Z", "+00:00"))

        if latest_date is None or commit_date > latest_date:
            latest_tag = tag["name"]
            latest_date = commit_date
    print(f"Latest tag {latest_tag} found. (Date: {latest_date})")
    return latest_tag


def get_filename_from_cd(cd_header):
    """Extracts filename from Content-Disposition header."""
    if not cd_header:
        return None
    fname_match = re.findall('filename="?([^"]+)"?', cd_header)
    if fname_match:
        return unquote(fname_match[0])
    return None

def download_projects(projects):
    """
    Downloads the latest source zip archive for each GitHub project from a list.
    Extracts the zip, deletes it, and records metadata for each project.
    """
    meta_data = []

    for project in projects:
        owner, repo = project.split('/')
        tag_name = None

        # First, try to get the latest release
        api_url = f"https://api.github.com/repos/{owner}/{repo}/releases/latest"
        response = robust_github_get(api_url, headers=GITHUB_HEADERS)

        if response.status_code == 200:
            release_data = response.json()
            tag_name = release_data.get('tag_name')

        # If no release, try to get latest tag by commit date
        if not tag_name:
            print(f"No release found for {project}, falling back to latest tag.")
            tag_name = get_latest_tag_by_commit_date(owner, repo)
            if not tag_name:
                print(f"Failed to get tags for {project}")
                meta_data.append({
                    "repoName": project,
                    "projectVersion": "N/A",
                    "projectName": repo,
                    "url": "NA",
                    "sourceRoot": "NA"
                })
                continue

        zip_url = f"https://github.com/{owner}/{repo}/archive/refs/tags/{tag_name}.zip"
        print(f"Downloading from {zip_url}")

        try:
            zip_response = robust_github_get(zip_url, headers=GITHUB_HEADERS)
            zip_response.raise_for_status()

            # Get actual filename from headers or sanitize fallback
            cd = zip_response.headers.get("Content-Disposition")
            actual_filename = get_filename_from_cd(cd)
            if not actual_filename:
                actual_filename = f"{repo}-{tag_name.replace('/', '-')}.zip"

            zip_path = os.path.join(DOWNLOAD_DIR, actual_filename)

            # Save the zip file to disk
            with open(zip_path, 'wb') as f:
                f.write(zip_response.content)

            print(f"Saved {actual_filename} to {zip_path}")

            # Extract the zip file
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(DOWNLOAD_DIR)

            # Remove the zip after extraction
            os.remove(zip_path)

            # Get the extracted folder name
            extracted_dirs = [
                name for name in os.listdir(DOWNLOAD_DIR)
                if os.path.isdir(os.path.join(DOWNLOAD_DIR, name)) and name.startswith(repo)
            ]

            if not extracted_dirs:
                print(f"Could not find extracted directory for {actual_filename}")
                continue

            actual_project_name = extracted_dirs[0]
            extract_dir = os.path.join(DOWNLOAD_DIR, actual_project_name)

            # Store metadata
            meta_data.append({
                "repoName": project,
                "projectVersion": tag_name,
                "projectName": actual_project_name,
                "url": zip_url,
                "sourceRoot": extract_dir
            })

        except Exception as e:
            print(f"Failed to process {zip_url}: {e}")
    
    with open(os.path.join(OUTPUT_DIR, "projects_meta.json"), "w") as f:
        json.dump(meta_data, f, indent=2)
        
    return meta_data

def load_meta_data_from_file():
    with open(os.path.join(OUTPUT_DIR, "projects_meta.json"), "r") as f:
        return json.load(f)



def change_java_version(java_version):
    """
    Configures the environment to use a specific Java version.
    Works on both Windows and Linux/macOS.
    Returns a modified environment dict for subprocess calls.
    """
    env = os.environ.copy()
    system = platform.system()

    # Windows setup
    if system == "Windows":
        java_home = os.path.join("C:\\", "Program Files", "Java", f"jdk-{java_version}")
        java_bin = os.path.join(java_home, "bin")
        java_exec = os.path.join(java_bin, "java.exe")

    # Linux/macOS setup
    else:
        java_base_dir = "/usr/local/java"
        java_home = os.path.join(java_base_dir, f"jdk{java_version}")
        java_bin = os.path.join(java_home, "bin")
        java_exec = os.path.join(java_bin, "java")

        # Check if Java is already installed
        if not os.path.exists(os.path.join(java_bin, "javac")):
            print(f"Java {java_version} not found at {java_home}, attempting to download...")

            version_map = {
            "8":  ("jdk8u452-b09", "8u452b09"),
            "11": ("jdk-11.0.20+8", "11.0.20_8"),
            "17": ("jdk-17.0.8+7", "17.0.8_7"),
            "21": ("jdk-21.0.2+13", "21.0.2_13")
            }

            tag, release = version_map.get(str(java_version), (None, None))
            if not tag or not release:
                raise Exception(f"Unsupported Java version: {java_version}")

            archive_name = f"OpenJDK{java_version}U-jdk_x64_linux_hotspot_{release}.tar.gz"
            url = f"https://github.com/adoptium/temurin{java_version}-binaries/releases/download/{tag}/{archive_name}"


            try:
                os.makedirs(java_home, exist_ok=True)
                subprocess.run(
                    f'curl -Ls "{url}" | tar -xz -C "{java_home}" --strip-components=1',
                    shell=True,
                    check=True
                )
                print(f"Java {java_version} installed at {java_home}")
            except subprocess.CalledProcessError as e:
                raise Exception(f"Failed to download or extract Java {java_version}: {e}")

    # Update environment
    env["JAVA_HOME"] = java_home
    env["PATH"] = java_bin + os.pathsep + env.get("PATH", "")

    # Optional: verify java version
    try:
        result = subprocess.run([java_exec, "-version"], env=env, capture_output=True, text=True)
        print(f"Java version set to {java_version}:")
        print(result.stderr.strip() or result.stdout.strip())
    except Exception as e:
        print(f"Failed to verify Java version: {e}")

    return env


def create_codeql_database(source_root, db_name, env):
    """
    Attempts to create a CodeQL database for the given source root using the provided environment.
    Also verifies that Java 21 is used by both JAVA_HOME and PATH.
    """
    # print(f"\nVerifying Java version for {db_name}...")

    # # Check direct JAVA_HOME java.exe
    # java_exec = os.path.join(env["JAVA_HOME"], "bin", "java.exe")
    # java_check = subprocess.run([java_exec, "-version"], env=env, capture_output=True, text=True)
    # print("Explicit java version from JAVA_HOME:")
    # print(java_check.stderr.strip())

    # # Log first PATH entry
    # print("\nPATH being used by CodeQL:")
    # print(env["PATH"].split(os.pathsep)[0])

    # Build command
    command = [
        "codeql", "database", "create", db_name,
        "--language=java",
        "--overwrite",
        f"--source-root={source_root}"
    ]
    print(f"\nRunning CodeQL database create for {db_name}...\n")

    # Run command with real-time streaming output
    process = subprocess.Popen(command, env=env, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True)
    java_version_pattern = re.compile(r"Java version:.*", re.IGNORECASE)


    # for line in process.stdout:
    #     sys.stdout.write(line)
    #     sys.stdout.flush()

    for line in process.stdout:
        if java_version_pattern.search(line):
            print(line.strip())  # Only print the matching Java version line

    process.wait()

    return process.returncode == 0

def write_json(data):
    """
    Writes only successfully built project metadata (i.e., not 'DNB') to a JSON file.
    This output file is used by setup_projects.py to give PIQUE all the projects that build successfully.
    Output file: projects_auto.json
    """
    successful_projects = [proj for proj in data if proj["javaVersion"] != "DNB"]

    output = {"projects": successful_projects}
    output_path = os.path.join(OUTPUT_DIR, "projects_auto.json")

    with open(output_path, "w") as f:
        json.dump(output, f, indent=2)

    print(f"\nSaved successful builds to JSON: {output_path}")


def write_xlsx(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Build Results"

    headers = [
        "repoName", "projectVersion", "javaVersion", "projectName", "Would Build",
        "Accuracy (TP / Total)", "Accuracy With Validation", "Download Link"
    ]
    ws.append(headers)

    # Define fill colors
    green_fill = PatternFill(start_color="bfd6ac", end_color="bfd6ac", fill_type="solid")  # light green
    red_fill = PatternFill(start_color="d69c9b", end_color="d69c9b", fill_type="solid")    # light red
    orange_fill = PatternFill(start_color="f9cb9c", end_color="f9cb9c", fill_type="solid") # light orange

    for project in data:
        would_build = "No" if project["javaVersion"] == "DNB" else "Yes"
        row = [
            project["repoName"],
            project["projectVersion"],
            project["javaVersion"],
            project["projectName"],
            would_build,
            "",  # Accuracy (TP / Total)
            "",  # Accuracy With Validation
            project["url"]
        ]
        ws.append(row)

        # Apply color fill
        if project["url"] == "NA":
            fill = orange_fill
        elif would_build == "No":
            fill = red_fill
        else:
            fill = green_fill

        for cell in ws[ws.max_row]:
            cell.fill = fill

    output_file = os.path.join(OUTPUT_DIR, "projects_auto.xlsx")
    wb.save(output_file)
    print(f"Saved XLSX to {output_file}")



def build_project(project):
    source_root = project["sourceRoot"]

    curr_project_metadata = {
        "repoName": project["repoName"],
        "projectVersion": project["projectVersion"],
        "javaVersion": "DNB",
        "projectName": project["projectName"],
        "url": project["url"]
    }

    if project["url"] == "NA":
        print(f"Skipping {project['projectName']} due to missing URL.")
        return curr_project_metadata


    for java_version in java_versions:
        env = change_java_version(java_version)
        db_output = os.path.join(DATABASE_DIR, f"{project['projectName']}_db")

        if create_codeql_database(source_root, db_output, env):
            curr_project_metadata["javaVersion"] = java_version
            print(f"Successfully built {project['projectName']} with Java {java_version}")
            return curr_project_metadata

    # print(f"Failed to build {project['projectName']} with any Java version.")
    return curr_project_metadata


def main():
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    os.makedirs(DATABASE_DIR, exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    projects = read_projects(input_projects)
    project_index = {p: i for i, p in enumerate(projects)}

    # meta_data = download_projects(projects)
    meta_data = load_meta_data_from_file()
    project_results = []

    total_projects = len(meta_data)
    completed_projects = 0
    with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
        futures = [executor.submit(build_project, project) for project in meta_data]
        for future in concurrent.futures.as_completed(futures):
            result = future.result()
            project_results.append(result)
            completed_projects += 1
            print(f"\rCompleted {completed_projects}/{total_projects} projects.", end="")
            

    # ? Ensure original order from input
    project_results.sort(key=lambda x: project_index.get(x["repoName"], float("inf")))

    write_json(project_results)
    write_xlsx(project_results)

if __name__ == "__main__":
    main()
